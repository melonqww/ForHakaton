import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.cell import WriteOnlyCell, Cell
import pandas as pd

from src.config import CATEGORIES
from src.utils import find_column_index

class StreamingExcelExporter:
    def __init__(self, output_path: str, headers: list, text_col: int, group_col: int, rank_col: int, summary_col: int, type_col: int):
        self.output_path = output_path
        self.headers = headers
        self.text_col = text_col + 1 if text_col is not None else None
        self.group_col = group_col + 1 if group_col is not None else None
        self.rank_col = rank_col + 1 if rank_col is not None else None
        self.summary_col = summary_col + 1 if summary_col is not None else None
        self.type_col = type_col + 1 if type_col is not None else None
        
        self.wb = openpyxl.Workbook(write_only=True)
        self.ws_summary = self.wb.create_sheet("Сводка")
        self.ws_summary.views.sheetView[0].showGridLines = True
        self.ws = self.wb.create_sheet("Данные")
        self.ws.views.sheetView[0].showGridLines = True
        
        # Стилизация шапки
        header_fill = PatternFill(start_color="2B4C7E", end_color="2B4C7E", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        header_row = []
        for col_name in headers:
            cell = WriteOnlyCell(self.ws, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            header_row.append(cell)
            
        self.ws.append(header_row)
        self.row_count = 1
        
        # Сбор ширины столбцов по заголовкам
        self.col_widths = [0] * len(headers)
        for i, col_name in enumerate(headers):
            self.col_widths[i] = len(str(col_name))
            
    def write_row(self, row_data: list):
        cleaned_row = []
        for val in row_data:
            if pd.isna(val):
                cleaned_row.append("")
            elif hasattr(val, 'strftime'):
                cleaned_row.append(val.strftime('%Y-%m-%d %H:%M:%S'))
            else:
                cleaned_row.append(val)
                
        self.ws.append(cleaned_row)
        self.row_count += 1
        
        # Сбор ширины по первым 1000 строк
        if self.row_count <= 1000:
            for i, val in enumerate(cleaned_row):
                if i < len(self.col_widths):
                    val_str = str(val or '')
                    if len(val_str) > self.col_widths[i]:
                        self.col_widths[i] = len(val_str)
                        
    def close(self, stats: dict):
        # Наложение ширины колонок
        for col_idx in range(1, len(self.col_widths) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = self.col_widths[col_idx - 1]
            
            # Увеличенная ширина для текстовых полей
            if col_idx == self.text_col or col_idx == self.summary_col or (col_idx <= len(self.headers) and self.headers[col_idx-1] == "Очищенный текст"):
                self.ws.column_dimensions[col_letter].width = 40
            else:
                self.ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 25)
                
        # Выпадающие списки для групп категорий
        if self.group_col:
            group_letter = get_column_letter(self.group_col)
            dv = DataValidation(
                type="list", 
                formula1=f'"{",".join(CATEGORIES)}"', 
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Ошибка ввода",
                error="Пожалуйста, выберите категорию из списка утвержденных Минцифры."
            )
            self.ws.data_validations.append(dv)
            dv.add(f"{group_letter}2:{group_letter}{self.row_count}")
            
        # Цветовое шкалирование важности (градиент)
        if self.rank_col:
            rank_col_letter = get_column_letter(self.rank_col)
            color_scale = ColorScaleRule(
                start_type='num', start_value=1, start_color='E2F0D9',
                mid_type='num', mid_value=3, mid_color='FFF2CC',
                end_type='num', end_value=5, end_color='FCE4D6'
            )
            self.ws.conditional_formatting.add(f"{rank_col_letter}2:{rank_col_letter}{self.row_count}", color_scale)
            
        # Использование ранее созданного листа аналитической сводки
        ws_summary = self.ws_summary
        
        title_font = Font(name='Calibri', size=14, bold=True, color='2B4C7E')
        section_font = Font(name='Calibri', size=12, bold=True)
        bold_font = Font(name='Calibri', size=11, bold=True)
        normal_font = Font(name='Calibri', size=11)
        header_fill = PatternFill(start_color="2B4C7E", end_color="2B4C7E", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")
        center_align = Alignment(horizontal="center")
        wrap_align = Alignment(wrap_text=True)
        
        def write_styled_row(ws, cells_data):
            row_cells = []
            for item in cells_data:
                if isinstance(item, Cell):
                    row_cells.append(item)
                else:
                    val = item[0] if isinstance(item, tuple) else item
                    cell = WriteOnlyCell(ws, value=val)
                    if isinstance(item, tuple):
                        if len(item) > 1 and item[1]: cell.font = item[1]
                        if len(item) > 2 and item[2]: cell.fill = item[2]
                        if len(item) > 3 and item[3]: cell.alignment = item[3]
                    row_cells.append(cell)
            ws.append(row_cells)
            
        # Заголовок
        write_styled_row(ws_summary, [("Сводный анализ обращений граждан", title_font)])
        write_styled_row(ws_summary, [])
        
        # Общие счетчики
        total = stats.get("total_count", 0)
        problems = stats.get("problems_count", 0)
        write_styled_row(ws_summary, [("Всего обращений:", bold_font), (total, normal_font)])
        write_styled_row(ws_summary, [("Реальных проблем:", bold_font), (problems, normal_font)])
        write_styled_row(ws_summary, [("Спам/благодарности:", bold_font), (total - problems, normal_font)])
        write_styled_row(ws_summary, [])
        
        # ТОП-3 района
        write_styled_row(ws_summary, [("ТОП-3 проблемных муниципалитета", section_font)])
        
        headers_top3 = ['Район', 'Обращений', 'Основная категория', 'Ср. критичность', 'Критичных (4-5)', 'Ключевые проблемы']
        top3_header_cells = []
        for h in headers_top3:
            cell = WriteOnlyCell(ws_summary, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            top3_header_cells.append(cell)
        ws_summary.append(top3_header_cells)
        
        for item in stats.get("top3_districts", []):
            write_styled_row(ws_summary, [
                (item["district"], bold_font),
                (item["count"], normal_font, None, center_align),
                (item["top_cat"], normal_font),
                (round(item["avg_rank"], 1), normal_font, None, center_align),
                (item["critical_count"], normal_font, None, center_align),
                (item["key_problems"], normal_font, None, wrap_align)
            ])
            
        write_styled_row(ws_summary, [])
        
        # ТОП-10 районов
        write_styled_row(ws_summary, [("ТОП-10 районов по количеству обращений", section_font)])
        
        headers_top10 = ['Район', 'Обращений', 'Основная категория', 'Ср. критичность']
        top10_header_cells = []
        for h in headers_top10:
            cell = WriteOnlyCell(ws_summary, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            top10_header_cells.append(cell)
        ws_summary.append(top10_header_cells)
        
        for item in stats.get("top10_districts", []):
            write_styled_row(ws_summary, [
                (item["district"], normal_font),
                (item["count"], normal_font, None, center_align),
                (item["top_cat"], normal_font),
                (round(item["avg_rank"], 1), normal_font, None, center_align)
            ])
            
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 14
        ws_summary.column_dimensions['C'].width = 25
        ws_summary.column_dimensions['D'].width = 18
        ws_summary.column_dimensions['E'].width = 18
        ws_summary.column_dimensions['F'].width = 60
        
        self.wb.save(self.output_path)
        self.wb.close()

def export_to_excel(df: pd.DataFrame, output_path: str):
    """Метод-обертка для совместимости с DataFrame-ориентированными тестами."""
    headers = list(df.columns)
    
    text_col = find_column_index(df, "text", 36)
    group_col = find_column_index(df, "group", 21)
    
    cols_list = list(df.columns)
    rank_col = cols_list.index("Ранг критичности") if "Ранг критичности" in cols_list else None
    summary_col = cols_list.index("Краткое саммари") if "Краткое саммари" in cols_list else None
    type_col = cols_list.index("Тип инцидента") if "Тип инцидента" in cols_list else None
    
    df_problems = df[df['Тип инцидента'] == 'Проблема'] if 'Тип инцидента' in df.columns else df
    
    stats = {
        "total_count": len(df),
        "problems_count": len(df_problems)
    }
    
    if not df_problems.empty and 'Нормализованное Гео' in df.columns:
        group_col_name = df.columns[find_column_index(df, 'group', 21)]
        
        # Сбор ТОП-3
        top3_districts = []
        top3 = df_problems['Нормализованное Гео'].value_counts().head(3)
        for district, count in top3.items():
            d_df = df_problems[df_problems['Нормализованное Гео'] == district]
            cats = d_df[group_col_name].value_counts()
            top_cat = cats.index[0] if not cats.empty else ''
            avg_rank = d_df['Ранг критичности'].mean() if 'Ранг критичности' in d_df.columns else 0
            critical_count = len(d_df[d_df['Ранг критичности'] >= 4]) if 'Ранг критичности' in d_df.columns else 0
            
            key_problems = ''
            if 'Краткое саммари' in d_df.columns and 'Ранг критичности' in d_df.columns:
                critical = d_df[d_df['Ранг критичности'] >= 4]
                if not critical.empty:
                    summaries = critical['Краткое саммари'].head(3).tolist()
                    key_problems = '; '.join([s for s in summaries if s])
                    
            top3_districts.append({
                "district": district,
                "count": count,
                "top_cat": top_cat,
                "avg_rank": avg_rank,
                "critical_count": critical_count,
                "key_problems": key_problems
            })
        stats["top3_districts"] = top3_districts
        
        # Сбор ТОП-10
        top10_districts = []
        top10 = df_problems['Нормализованное Гео'].value_counts().head(10)
        for district, count in top10.items():
            d_df = df_problems[df_problems['Нормализованное Гео'] == district]
            cats = d_df[group_col_name].value_counts()
            top_cat = cats.index[0] if not cats.empty else ''
            avg_rank = d_df['Ранг критичности'].mean() if 'Ранг критичности' in d_df.columns else 0
            
            top10_districts.append({
                "district": district,
                "count": count,
                "top_cat": top_cat,
                "avg_rank": avg_rank
            })
        stats["top10_districts"] = top10_districts
        
    exporter = StreamingExcelExporter(output_path, headers, text_col, group_col, rank_col, summary_col, type_col)
    for row in df.values.tolist():
        exporter.write_row(row)
    exporter.close(stats)
